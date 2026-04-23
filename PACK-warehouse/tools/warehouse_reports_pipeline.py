#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Складской pipeline:
1. Находит свежие warehouse CSV после sync Google Sheets.
2. Создает карточку по каждому отчету.
3. Обновляет сводный markdown-отчет.
4. Опционально отправляет сводку в Telegram.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
import time
from functools import lru_cache
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable
from urllib.parse import quote, urlencode
from urllib.request import Request, urlopen


def repo_root() -> Path:
    # .../VK-offee/PACK-warehouse/tools/warehouse_reports_pipeline.py
    return Path(__file__).resolve().parents[2]


ROOT = repo_root()
KB_DIR = ROOT / "knowledge-base"
WAREHOUSE_DIR = ROOT / "PACK-warehouse"
CARDS_DIR = WAREHOUSE_DIR / "02-domain-entities" / "report-cards"
WP_DIR = WAREHOUSE_DIR / "04-work-products"
LATEST_REPORT = WP_DIR / "WH.REPORT.002-warehouse-sync-summary-latest.md"
DECISION_QUEUE = WP_DIR / "WH.SESSION.001-decision-queue-latest.md"
REGISTRY_FILE = WP_DIR / "WH.REGISTRY.001-documents.csv"
DLQ_DIR = WAREHOUSE_DIR / "03-quarantine" / "dlq-files"
DLQ_REPORT = WP_DIR / "WH.DLQ.001-quarantine-report.md"
BOT_REPORTS_DIR = KB_DIR / "Отчёты для бота" / "Склад"
SUPPLIER_DIRECTORY_FILE = WAREHOUSE_DIR / "02-domain-entities" / "WH.SUPPLIER.001-directory.md"
PROCUREMENT_REPORT_FILE = WAREHOUSE_DIR / "04-work-products" / "WH.WP.007-invoice-procurement-supplier-map-2026-04-19 (Карта поставщиков и закупочный контур по PDF-накладным).md"

KEYWORDS = (
    "Остатки по складам",
    "Отчет по инвентаризации",
    "Продажи_",
    "Выручка_",
    "Каталог_",
    "Накладные_",
    "Продажи ",
    "Выручка ",
    "Каталог ",
    "Накладные ",
    "Список зерна Самокиша - Тургенева - Луговая - Склад",
    "Список зерна Самокиша - Тургенева - Луговая - Комментарий",
    "АБС анализ",
    "АБС-анализ",
    "ABC анализ",
    "ABC-анализ",
    "Abc",
    "АБС",
)

def sanitize_slug(text: str) -> str:
    out = text.lower()
    out = re.sub(r"[^a-zа-я0-9]+", "-", out)
    out = out.strip("-")
    return out[:120] or "report"


def detect_report_type(name: str) -> str:
    low = name.lower()
    if "продаж" in low:
        return "sales"
    if "выруч" in low:
        return "revenue"
    if "каталог" in low:
        return "catalog"
    if "накладн" in low:
        return "invoice"
    if ("комментарий" in low) or ("комментари" in low):
        return "comment"
    if "инвентаризац" in low:
        return "inventory"
    if "зерна" in low:
        return "beans"
    if "остатки" in low:
        return "stock"
    if ("абс" in low) or ("abc" in low):
        return "abc"
    return "other"


def file_sha1(path: Path) -> str:
    sha = hashlib.sha1()
    with path.open("rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            sha.update(chunk)
    return sha.hexdigest()


def now_stamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def escape_html(text: str) -> str:
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def normalize_ws(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "")).strip()


def money_to_float(raw: str) -> float:
    return float((raw or "").replace("\u00a0", "").replace(" ", "").replace(",", "."))


def normalize_invoice_item_name(text: str) -> str:
    value = normalize_ws(text)
    value = re.sub(
        r"^(?:1а\s+1б\s+2\s+2а\s+3\s+4\s+5\s+6\s+7\s+8\s+9\s+10\s+10а\s+11\s+)+",
        "",
        value,
        flags=re.IGNORECASE,
    )
    value = re.sub(r"^[A-ZА-Я0-9._/-]{2,}\s+1\s+", "", value)
    value = re.sub(r"^[A-ZА-Я0-9._/-]{2,}\s+", "", value)
    value = re.sub(r"\s{2,}", " ", value).strip(" ,.-")
    return value or normalize_ws(text)


def canonical_invoice_supplier_name(seller: str) -> str:
    seller_low = normalize_ws(seller).lower()
    if "барсервис" in seller_low:
        return "Барсервис"
    if "мфуд" in seller_low:
        return "МФУД"
    if "тейсти кофе" in seller_low or "тэйсти кофе" in seller_low or "tasty coffee" in seller_low:
        return "Тэйсти Кофе"
    if "unicava" in seller_low:
        return "UNICAVA"
    if "шипилов" in seller_low:
        return "ИП Шипилов Сергей Николаевич"
    if "камелот" in seller_low:
        return "КАМЕЛОТ"
    if "ритейл проперти 6" in seller_low:
        return "Ритейл Проперти 6"
    if "премиум-сегмент" in seller_low:
        return "Премиум-Сегмент"
    if "новая жизнь" in seller_low:
        return "Новая Жизнь"
    return normalize_ws(seller)


def parse_ru_date(raw: str) -> datetime | None:
    raw = normalize_ws(raw)
    months = {
        "января": "01", "февраля": "02", "марта": "03", "апреля": "04",
        "мая": "05", "июня": "06", "июля": "07", "августа": "08",
        "сентября": "09", "октября": "10", "ноября": "11", "декабря": "12",
    }
    word_date = re.search(r"(\d{1,2})\s+([а-я]+)\s+(\d{4})", raw.lower())
    if word_date and word_date.group(2) in months:
        raw = f"{int(word_date.group(1)):02d}.{months[word_date.group(2)]}.{word_date.group(3)}"
    for fmt in ("%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def canonical_pdf_key(pdf: Path) -> str:
    name = pdf.name.lower()
    while name.endswith(".pdf"):
        name = name[:-4]
    return name.strip()


def split_invoice_blocks(text: str) -> list[str]:
    parts = re.split(r"(?=передаточный\s+Счет-фактура №)", text, flags=re.IGNORECASE)
    return [p for p in parts if "Счет-фактура №" in p]


def parse_invoice_line_items(block: str, seller: str) -> list[dict[str, object]]:
    section_match = re.search(r"Наименование товара.*?Всего к оплате", block, flags=re.S)
    if not section_match:
        return []
    section = section_match.group(0)
    pattern = re.compile(
        r"^\s*(?P<code>\S+)\s+(?P<idx>\d+)\s+(?P<name>.+?)\s+-\s+\d+\s+\S+\s+"
        r"(?P<qty>\d+(?:[.,]\d+)?)\s+(?P<price>\d[\d \u00a0]*[.,]\d{2})\s+"
        r"(?P<amount>\d[\d \u00a0]*[.,]\d{2})",
        flags=re.M | re.S,
    )
    items: list[dict[str, object]] = []
    for match in pattern.finditer(section):
        name = normalize_invoice_item_name(match.group("name"))
        if len(name) < 3:
            continue
        try:
            qty = money_to_float(match.group("qty"))
            price = money_to_float(match.group("price"))
            amount = money_to_float(match.group("amount"))
        except ValueError:
            continue
        items.append(
            {
                "seller": seller,
                "code": match.group("code"),
                "name": name,
                "qty": qty,
                "price": price,
                "amount": amount,
            }
        )
    return items


@lru_cache(maxsize=1)
def collect_invoice_line_items() -> list[dict[str, object]]:
    pdfs = sorted(KB_DIR.rglob("*Накладные*.pdf"))
    unique_pdfs: list[Path] = []
    seen_keys: set[str] = set()
    for pdf in pdfs:
        key = canonical_pdf_key(pdf)
        if key in seen_keys:
            continue
        seen_keys.add(key)
        unique_pdfs.append(pdf)

    rows: list[dict[str, object]] = []
    for pdf in unique_pdfs:
        try:
            text = subprocess.check_output(["pdftotext", "-layout", str(pdf), "-"], text=True)
        except Exception:
            continue
        for block in split_invoice_blocks(text):
            no_date = re.search(r"Счет-фактура №\s*([^\s]+)\s+от\s+([^\n]+)", block)
            seller = re.search(r"Продавец\s+(.+)", block)
            if not (no_date and seller):
                continue
            seller_name = canonical_invoice_supplier_name(re.sub(r"\(\d+\)\s*$", "", seller.group(1)))
            invoice_no = no_date.group(1).strip()
            invoice_date_raw = normalize_ws(re.sub(r"\s+\(.*$", "", no_date.group(2)))
            invoice_date_dt = parse_ru_date(invoice_date_raw)
            for item in parse_invoice_line_items(block, seller_name):
                item["invoice_no"] = invoice_no
                item["invoice_date"] = invoice_date_raw
                item["invoice_dt"] = invoice_date_dt
                item["source_file"] = pdf.relative_to(ROOT).as_posix()
                rows.append(item)
    return rows


def detect_github_repo_web_url() -> str:
    try:
        remote = (
            subprocess.check_output(
                ["git", "-C", str(ROOT), "remote", "get-url", "origin"],
                text=True,
                stderr=subprocess.DEVNULL,
            )
            .strip()
            .rstrip("/")
        )
    except Exception:
        return ""

    if remote.startswith("git@github.com:"):
        path = remote.split("git@github.com:", 1)[1]
        if path.endswith(".git"):
            path = path[:-4]
        return f"https://github.com/{path}"
    if remote.startswith("https://github.com/"):
        if remote.endswith(".git"):
            remote = remote[:-4]
        return remote
    return ""


def detect_git_branch() -> str:
    try:
        branch = (
            subprocess.check_output(
                ["git", "-C", str(ROOT), "rev-parse", "--abbrev-ref", "HEAD"],
                text=True,
                stderr=subprocess.DEVNULL,
            )
            .strip()
        )
        if branch and branch != "HEAD":
            return branch
    except Exception:
        pass

    # Detached HEAD fallback: пробуем origin/HEAD -> origin/main
    try:
        ref = (
            subprocess.check_output(
                ["git", "-C", str(ROOT), "symbolic-ref", "refs/remotes/origin/HEAD"],
                text=True,
                stderr=subprocess.DEVNULL,
            )
            .strip()
        )
        if ref.startswith("refs/remotes/origin/"):
            cand = ref.split("refs/remotes/origin/", 1)[1].strip()
            if cand:
                return cand
    except Exception:
        pass

    return "main"


def path_exists_in_head(rel_path: str) -> bool:
    if not rel_path:
        return False
    try:
        subprocess.check_call(
            ["git", "-C", str(ROOT), "cat-file", "-e", f"HEAD:{rel_path}"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        return True
    except Exception:
        return False


def build_github_link(rel_path: str) -> str:
    base = (
        os.getenv("WAREHOUSE_REPORT_BASE_URL", "").strip()
        or detect_github_repo_web_url()
    )
    branch = os.getenv("WAREHOUSE_REPORT_BRANCH", "").strip() or detect_git_branch()
    if not base:
        return ""
    if not path_exists_in_head(rel_path):
        return ""
    return f"{base}/blob/{branch}/{quote(rel_path, safe='/')}"


def report_type_human(report_type: str) -> str:
    return {
        "stock": "Остатки",
        "inventory": "Инвентаризация",
        "inventory_non_stock": "Инвентаризация (сводная)",
        "beans": "Зерно",
        "abc": "ABC-анализ",
        "sales": "Продажи",
        "revenue": "Выручка",
        "catalog": "Каталог",
        "invoice": "Накладные",
        "comment": "Комментарии",
        "other": "Отчёт",
    }.get((report_type or "").strip(), "Отчёт")


def telegram_item_title(item: dict) -> str:
    source_rel = str(item.get("source_rel", "") or "")
    source_name = Path(source_rel).name if source_rel else str(item.get("title", "Отчёт"))
    source_name = re.sub(r"\.(csv|pdf)$", "", source_name, flags=re.IGNORECASE)
    source_name = re.sub(r"\s+-\s+лист\d+$", "", source_name, flags=re.IGNORECASE)
    source_name = re.sub(r"\s+-\s+таблица$", "", source_name, flags=re.IGNORECASE)
    source_name = source_name.replace("_", " ").strip()
    kind = report_type_human(str(item.get("report_type", "") or ""))
    return f"{kind}: {source_name}" if source_name else kind


def load_registry() -> dict[str, dict[str, str]]:
    if not REGISTRY_FILE.exists():
        return {}
    rows: dict[str, dict[str, str]] = {}
    with REGISTRY_FILE.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = (row.get("record_key") or "").strip()
            if key:
                rows[key] = row
    return rows


def save_registry(rows: dict[str, dict[str, str]]) -> None:
    REGISTRY_FILE.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "record_key",
        "source_file",
        "source_mtime",
        "source_size_bytes",
        "source_hash",
        "report_type",
        "status",
        "first_seen_at",
        "last_seen_at",
        "last_processed_at",
        "card_path",
        "bot_card_path",
        "dlq_path",
        "error",
    ]
    with REGISTRY_FILE.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for key in sorted(rows.keys()):
            writer.writerow(rows[key])


def find_recent_reports(hours: int) -> list[Path]:
    cutoff = datetime.now() - timedelta(hours=hours)
    items: list[Path] = []
    for ext in ("*.csv", "*.pdf", "*.xlsx"):
        for p in KB_DIR.rglob(ext):
            try:
                mtime = datetime.fromtimestamp(p.stat().st_mtime)
            except FileNotFoundError:
                continue
            if mtime < cutoff:
                continue
            name = p.name
            if any(k in name for k in KEYWORDS):
                items.append(p)
    return sorted(items)


def read_text_best_effort(path: Path) -> str:
    raw = path.read_bytes()
    candidates = ("utf-8-sig", "utf-8", "cp1251", "windows-1251", "latin-1")
    best_text = ""
    best_score = -1
    for enc in candidates:
        try:
            text = raw.decode(enc, errors="strict")
        except Exception:
            continue
        score = sum(1 for ch in text if ("а" <= ch.lower() <= "я")) - text.count("�")
        if score > best_score:
            best_text = text
            best_score = score
    if best_text:
        return best_text
    return raw.decode("utf-8", errors="ignore")


def parse_csv_rows(path: Path) -> list[list[str]]:
    raw = read_text_best_effort(path)
    sample = raw[:8000]
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except Exception:
        for cand in (";", "\t", "|", ","):
            if sample.count(cand) >= 3:
                delimiter = cand
                break
    rows = list(csv.reader(raw.splitlines(), delimiter=delimiter))
    if not rows:
        raise ValueError("empty csv")
    return rows


def parse_xlsx_rows(path: Path) -> list[list[str]]:
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(path, read_only=False, data_only=True)
    best_rows: list[list[str]] = []
    best_score = -1
    for ws in wb.worksheets:
        rows: list[list[str]] = []
        non_empty = 0
        for row in ws.iter_rows(values_only=True):
            vals = ["" if v is None else str(v).strip() for v in row]
            if any(vals):
                non_empty += 1
            rows.append(vals)
        if non_empty > best_score:
            best_score = non_empty
            best_rows = rows
    if not best_rows:
        raise ValueError("empty xlsx")
    return best_rows


def parse_xlsx_sheets(path: Path) -> list[dict[str, object]]:
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(path, read_only=False, data_only=True)
    sheets: list[dict[str, object]] = []
    for ws in wb.worksheets:
        rows: list[list[str]] = []
        non_empty = 0
        for row in ws.iter_rows(values_only=True):
            vals = ["" if v is None else str(v).strip() for v in row]
            if any(vals):
                non_empty += 1
            rows.append(vals)
        sheets.append(
            {
                "title": ws.title,
                "rows": rows,
                "non_empty": non_empty,
            }
        )
    if not sheets:
        raise ValueError("empty xlsx")
    return sheets


def parse_tabular_rows(path: Path) -> list[list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return parse_csv_rows(path)
    if suffix == ".xlsx":
        return parse_xlsx_rows(path)
    raise ValueError(f"unsupported tabular suffix: {suffix}")


def normalize_abc_class(value: str) -> str:
    cat = value.strip().upper()
    return cat.replace("А", "A").replace("В", "B").replace("С", "C")


def parse_abc_payload(rows: list[list[str]], *, sheet_name: str = "") -> dict[str, object]:
    """Извлечь ABC-категории и метаданные качества разбора."""
    best_rows: list[list[str]] = []
    result: dict[str, str] = {}
    unmatched: list[str] = []
    payload: dict[str, object] = {
        "sheet_name": sheet_name,
        "categories": result,
        "matched_rows": 0,
        "unmatched_rows": 0,
        "total_rows": 0,
        "header_row_idx": 0,
        "name_col": -1,
        "cat_col": -1,
        "status": "empty",
        "unmatched_preview": unmatched,
    }
    if not rows:
        return payload

    payload["total_rows"] = len(rows)

    header_row_idx = 0
    name_col = -1
    cat_col = -1
    scan_limit = min(len(rows), 16)
    for ridx in range(scan_limit):
        header = [c.strip().lower() for c in rows[ridx]]
        local_name = -1
        local_cat = -1
        for i, h in enumerate(header):
            if any(kw in h for kw in ("наимен", "товар", "позиц", "продукт", "name", "номенклат")) and local_name == -1:
                local_name = i
            if any(kw in h for kw in ("категор", "класс", "abc", "абс", "group")) and local_cat == -1:
                local_cat = i
        if local_name != -1 and local_cat != -1:
            header_row_idx = ridx
            name_col = local_name
            cat_col = local_cat
            break

    if name_col == -1:
        name_col = 0
    if cat_col == -1:
        width = max(len(r) for r in rows)
        for i in range(width):
            vals = {
                normalize_abc_class(r[i])
                for r in rows[header_row_idx + 1 :]
                if len(r) > i and r[i].strip()
            }
            if vals and vals.issubset({"A", "B", "C"}):
                cat_col = i
                break

    payload["header_row_idx"] = header_row_idx
    payload["name_col"] = name_col
    payload["cat_col"] = cat_col

    if cat_col == -1:
        payload["status"] = "missing_category_column"
        return payload

    for r in rows[header_row_idx + 1 :]:
        if len(r) <= max(name_col, cat_col):
            continue
        name = r[name_col].strip()
        cat = normalize_abc_class(r[cat_col])
        if not name:
            continue
        if cat in ("A", "B", "C"):
            result[name.lower()] = cat
        else:
            unmatched.append(name)

    payload["matched_rows"] = len(result)
    payload["unmatched_rows"] = len(unmatched)
    payload["status"] = "ok" if result else "no_valid_categories"
    payload["unmatched_preview"] = unmatched[:8]
    return payload


def extract_pdf_payload(path: Path) -> dict[str, object]:
    payload: dict[str, object] = {
        "text": "",
        "pages": 0,
        "method": "",
        "error": "",
    }

    # 1) Pure-Python parser (предпочтительно для portability).
    try:
        from pypdf import PdfReader  # type: ignore

        reader = PdfReader(str(path))
        texts: list[str] = []
        for page in reader.pages:
            try:
                texts.append((page.extract_text() or "").strip())
            except Exception:
                texts.append("")
        text = "\n".join(t for t in texts if t)
        if text.strip():
            payload["text"] = text
            payload["pages"] = len(reader.pages)
            payload["method"] = "pypdf"
            return payload
    except Exception as exc:
        payload["error"] = str(exc)

    # 2) Fallback to pdftotext CLI (если доступен в окружении).
    try:
        proc = subprocess.run(
            ["pdftotext", "-layout", str(path), "-"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=True,
            text=True,
        )
        text = (proc.stdout or "").strip()
        if text:
            payload["text"] = text
            payload["method"] = "pdftotext"
            return payload
    except Exception as exc:
        payload["error"] = f"{payload.get('error', '')}; {exc}".strip("; ")

    return payload


def parse_invoice_pdf_metrics(text: str) -> dict[str, object]:
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    money_raw = re.findall(r"\b\d{1,3}(?:[ \u00a0]\d{3})*(?:[.,]\d{2})\b", text)
    money_vals: list[float] = []
    for m in money_raw:
        val = m.replace("\u00a0", "").replace(" ", "").replace(",", ".")
        try:
            money_vals.append(float(val))
        except ValueError:
            continue

    invoice_numbers = re.findall(r"(?:накладн\w*\s*№?\s*|№\s*)(\d{4,})", text, flags=re.IGNORECASE)
    dates = re.findall(r"\b\d{2}[./]\d{2}[./]\d{4}\b", text)
    vendors: list[str] = []
    for ln in lines[:250]:
        low = ln.lower()
        if any(k in low for k in ("ип ", "ооо ", "зао ", "поставщик", "продавец")):
            vendors.append(ln)
    vendors = list(dict.fromkeys(vendors))[:5]

    return {
        "line_count": len(lines),
        "char_count": len(text),
        "money_count": len(money_vals),
        "money_top": sorted(money_vals, reverse=True)[:8],
        "invoice_numbers": sorted(set(invoice_numbers))[:20],
        "dates": sorted(set(dates))[:10],
        "vendors": vendors,
        "preview_lines": lines[:12],
    }


def parse_stock_metrics(rows: list[list[str]]) -> dict:
    # Ожидаем формат: Наименование,Общее кол-во,Тургенева,Луговая,Самокиша
    metrics = {
        "rows": 0,
        "low_stock_count": 0,
        "low_stock_items": [],
        "item_totals": {},
        "totals_by_location": {"Тургенева": 0, "Луговая": 0, "Самокиша": 0},
    }
    if not rows:
        return metrics

    for r in rows[1:]:
        if not r or not r[0].strip():
            continue
        row = r + [""] * (5 - len(r))
        name = row[0].strip()
        if name.count("?") >= 3:
            # Поврежденная кодировка в legacy CSV не должна попадать в решение руководителя.
            continue
        total_raw = row[1].strip()
        if not total_raw or not re.fullmatch(r"\d+(\.\d+)?", total_raw):
            continue
        total = int(float(total_raw))
        metrics["rows"] += 1
        metrics["item_totals"][normalize_item_name(name)] = total

        for idx, loc in ((2, "Тургенева"), (3, "Луговая"), (4, "Самокиша")):
            cell = row[idx].strip()
            if cell.isdigit():
                metrics["totals_by_location"][loc] += int(cell)

        if total <= 3:
            metrics["low_stock_count"] += 1
            metrics["low_stock_items"].append((name, total))

    return metrics


def parse_comment_bullets(rows: list[list[str]], max_lines: int = 12) -> list[str]:
    text_chunks: list[str] = []
    for r in rows:
        for c in r:
            cell = c.strip()
            if cell:
                text_chunks.append(cell)
    merged = "\n".join(text_chunks)
    lines = [ln.strip() for ln in merged.splitlines() if ln.strip()]
    # убираем слишком шумные строки
    lines = [ln for ln in lines if len(ln) >= 3]
    return lines[:max_lines]


def parse_abc_categories(rows: list[list[str]]) -> dict[str, str]:
    """Извлечь маппинг {название_позиции: категория A/B/C} из ABC-анализа."""
    return dict(parse_abc_payload(rows).get("categories") or {})


def parse_number(value: str) -> float | None:
    raw = value.strip().replace("\u00a0", "").replace(" ", "")
    if not raw:
        return None
    raw = raw.replace(",", ".")
    if not re.fullmatch(r"-?\d+(\.\d+)?", raw):
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def extract_period_from_text(text: str) -> str:
    dates = re.findall(r"\d{2}\.\d{2}\.\d{4}", text)
    if len(dates) >= 2:
        return f"{dates[0]} — {dates[1]}"
    dates_iso = re.findall(r"\d{4}-\d{2}-\d{2}", text)
    if len(dates_iso) >= 2:
        return f"{dates_iso[0]} — {dates_iso[1]}"
    return "n/a"


def parse_table_profile(rows: list[list[str]]) -> dict[str, object]:
    max_cols = max((len(r) for r in rows), default=0)
    data_rows = rows[1:] if len(rows) > 1 else []
    non_empty_rows = sum(1 for r in data_rows if any(c.strip() for c in r))
    headers = [c.strip() for c in (rows[0] if rows else []) if c.strip()]

    numeric_col_stats: list[tuple[int, str, float]] = []
    if rows:
        width = len(rows[0])
        for i in range(width):
            vals = []
            for r in data_rows:
                if len(r) <= i:
                    continue
                num = parse_number(r[i])
                if num is not None:
                    vals.append(num)
            if vals:
                label = rows[0][i].strip() if i < len(rows[0]) else f"col_{i+1}"
                numeric_col_stats.append((i, label or f"col_{i+1}", sum(vals)))
    numeric_col_stats.sort(key=lambda x: abs(x[2]), reverse=True)

    return {
        "rows_total": len(rows),
        "rows_data": non_empty_rows,
        "cols": max_cols,
        "headers": headers[:8],
        "top_numeric": numeric_col_stats[:3],
    }


def parse_non_stock_inventory_metrics(rows: list[list[str]]) -> dict[str, object]:
    names: list[str] = []
    debt_notes: list[str] = []
    money_vals: list[float] = []

    money_re = re.compile(r"\d{1,3}(?:[ \u00a0]\d{3})*(?:[.,]\d{2})\s*₽?")
    for r in rows:
        if len(r) > 1:
            candidate = (r[1] or "").strip()
            if candidate and not re.search(r"\d", candidate) and len(candidate) >= 3:
                names.append(candidate)
        for c in r:
            cell = (c or "").strip()
            if not cell:
                continue
            if "долг" in cell.lower():
                debt_notes.append(cell)
            for m in money_re.findall(cell):
                raw = m.replace("₽", "").replace("\u00a0", "").replace(" ", "").replace(",", ".").strip()
                try:
                    money_vals.append(float(raw))
                except ValueError:
                    continue
    names = list(dict.fromkeys(names))
    debt_notes = list(dict.fromkeys(debt_notes))
    return {
        "people_count": len(names),
        "people_top": names[:10],
        "money_cells_count": len(money_vals),
        "money_cells_sum": round(sum(money_vals), 2) if money_vals else 0.0,
        "debt_notes": debt_notes[:6],
    }


def normalize_item_name(value: str) -> str:
    s = (value or "").lower().strip()
    s = s.replace("ё", "е")
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^a-zа-я0-9 ]+", "", s)
    return s.strip()


def normalize_item_family(value: str) -> str:
    s = normalize_item_name(value)
    s = s.replace("капельница", "дрип")
    s = s.replace("капельный пакет", "дрип")
    s = s.replace("дрип пакеты", "дрип")
    s = s.replace("дриппакеты", "дрип")
    s = s.replace("дрип пакет", "дрип")
    s = s.replace("субмарина дрипы", "")
    s = re.sub(r"\bдрип\b", "", s)
    s = re.sub(r"\b(10|12|16|30|50|100|200)\s*шт\b", "", s)
    s = re.sub(r"\b(100|200|250|500|1000)\s*гр\b", "", s)
    s = re.sub(r"\b1\s*кг\b", "", s)
    s = re.sub(r"\b1кг\b", "", s)
    s = re.sub(r"\b250\b", "", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def canonical_item_label(value: str) -> str:
    text = (value or "").strip()
    low = text.lower()
    low = low.replace("капельница", "дрип")
    low = low.replace("капельный пакет", "дрип")
    low = re.sub(r"\s+", " ", low).strip()
    return low.title()


def resolve_abc_category(item_name: str, abc_map: dict[str, str]) -> tuple[str | None, str]:
    direct = abc_map.get(normalize_item_name(item_name))
    if direct:
        return direct, "direct"

    family = normalize_item_family(item_name)
    if not family:
        return None, "missing"

    family_matches = {
        cat for name_norm, cat in abc_map.items()
        if normalize_item_family(name_norm) == family
    }
    if len(family_matches) == 1:
        return next(iter(family_matches)), "family"
    if len(family_matches) > 1:
        return None, "ambiguous_family"
    return None, "missing"


def classify_item_product_type(item_name: str) -> str:
    low = normalize_item_name(item_name)
    if any(k in low for k in ("вупи", "чизкейк", "моти", "шу", "десерт", "пирож", "эклер")):
        return "десерты"
    if any(k in low for k in ("блин", "ролл", "киш", "боул", "сэндвич", "онигур", "онигуру", "удон", "суп", "сырник", "чиабат", "салат", "гречк", "рис ", "соба", "том ям")):
        return "кухня"
    if any(k in low for k in ("дрип",)):
        return "кофе_drip"
    if any(k in low for k in ("плитка", "unicava", "какао классический")):
        return "шоколад"
    if any(k in low for k in ("зерно", "эспрессо", "кофе")):
        return "кофе_зерно"
    if re.search(r"\b(250|1000)\b", low) or "1кг" in low or "1 кг" in low:
        if any(
            marker in low
            for marker in (
                "бэрри", "фрутти", "колумбия", "эфиопия", "руанда", "коста рика",
                "гватемала", "бразилия", "богота", "гуджи", "кигали", "мукера",
                "декаф", "уила", "сидамо", "киву", "сан хосе", "сантьяго",
                "каранави", "маунт", "иргачефф", "оромия", "челбеса", "гейша",
            )
        ):
            return "кофе_зерно"
    if "шоколад" in low:
        return "шоколад"
    if "чай" in low:
        return "чай"
    if any(k in low for k in ("сироп",)):
        return "сиропы"
    if any(k in low for k in ("моющ", "хоз", "чист", "салфет", "перчат", "крышк", "стакан")):
        return "расходники"
    return "другое"


@lru_cache(maxsize=1)
def load_supplier_directory() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    if not SUPPLIER_DIRECTORY_FILE.exists():
        return rows
    text = SUPPLIER_DIRECTORY_FILE.read_text(encoding="utf-8", errors="ignore")
    lines = text.splitlines()
    header: list[str] | None = None
    in_table = False
    for line in lines:
        ln = line.strip()
        if not in_table:
            if ln.startswith("| supplier_name |"):
                parts = [p.strip() for p in ln.strip("|").split("|")]
                if parts:
                    header = parts
                    in_table = True
                continue
            continue
        if not ln.startswith("|"):
            break
        parts = [p.strip() for p in ln.strip("|").split("|")]
        if not parts:
            continue
        if set("".join(parts)) <= {"-", ":"}:
            continue
        if header and len(parts) >= len(header):
            row = {header[i]: parts[i] for i in range(len(header))}
            rows.append(row)
    return rows


def infer_supplier_for_item(item_name: str) -> dict[str, str]:
    product_type = classify_item_product_type(item_name)
    item_low = normalize_item_name(item_name)
    suppliers = load_supplier_directory()
    type_fragments = {
        "кофе_drip": ("кофе_drip", "кофе"),
        "кофе_зерно": ("кофе_зерно", "кофе"),
        "чай": ("чай",),
        "шоколад": ("шоколад",),
        "сиропы": ("сироп",),
        "расходники": ("расходник", "хоз"),
        "десерты": ("десерт",),
    }
    supplier_override_by_item = {
        "субмарина": "Субмарина",
        "unicava": "UNICAVA",
        "плитка": "UNICAVA",
        "бэрри 250": "Тэйсти Кофе",
        "фрутти 250": "Тэйсти Кофе",
        "фрутти 1кг": "Тэйсти Кофе",
        "бэрри 1кг": "Тэйсти Кофе",
        "колумбия декаф": "Тэйсти Кофе",
        "руанда кигали": "Тэйсти Кофе",
        "коста рика сан хосе": "Тэйсти Кофе",
        "эфиопия мукера": "Тэйсти Кофе",
        "эфиопия гуджи": "Тэйсти Кофе",
        "колумбия богота": "Тэйсти Кофе",
        "компетишн колумбия": "Тэйсти Кофе",
    }
    for marker, supplier_name_override in supplier_override_by_item.items():
        if marker in item_low:
            for row in suppliers:
                if (row.get("supplier_name") or "").strip().lower() == supplier_name_override.lower():
                    return {
                        "supplier_name": row.get("supplier_name", supplier_name_override),
                        "supplier_contact": row.get("supplier_contact", "TBD"),
                        "order_channel": row.get("order_channel", "TBD"),
                        "order_cutoff_time": row.get("order_cutoff_time", "TBD"),
                        "typical_lead_time_days": row.get("typical_lead_time_days", "TBD"),
                        "product_type": product_type,
                    }
    for row in suppliers:
        product_types = (row.get("product_types") or "").lower()
        if product_type and product_type.lower() in product_types:
            return {
                "supplier_name": row.get("supplier_name", "Уточнить у Жанны"),
                "supplier_contact": row.get("supplier_contact", "TBD"),
                "order_channel": row.get("order_channel", "TBD"),
                "order_cutoff_time": row.get("order_cutoff_time", "TBD"),
                "typical_lead_time_days": row.get("typical_lead_time_days", "TBD"),
                "product_type": product_type,
            }
        if product_type in type_fragments and any(frag in product_types for frag in type_fragments[product_type]):
            return {
                "supplier_name": row.get("supplier_name", "Уточнить у Жанны"),
                "supplier_contact": row.get("supplier_contact", "TBD"),
                "order_channel": row.get("order_channel", "TBD"),
                "order_cutoff_time": row.get("order_cutoff_time", "TBD"),
                "typical_lead_time_days": row.get("typical_lead_time_days", "TBD"),
                "product_type": product_type,
            }
    fallback_by_type = {
        "кофе_drip": "Тэйсти Кофе",
        "кофе_зерно": "Тэйсти Кофе",
        "чай": "Тэйсти Кофе",
        "шоколад": "UNICAVA",
        "сиропы": "Барсервис",
        "десерты": "Дмитрий (Десерты)",
        "кухня": "Уточнить у Жанны",
    }
    fallback_name = fallback_by_type.get(product_type, "Уточнить у Жанны")
    for row in suppliers:
        if (row.get("supplier_name") or "").strip().lower() == fallback_name.lower():
            return {
                "supplier_name": row.get("supplier_name", fallback_name),
                "supplier_contact": row.get("supplier_contact", "TBD"),
                "order_channel": row.get("order_channel", "TBD"),
                "order_cutoff_time": row.get("order_cutoff_time", "TBD"),
                "typical_lead_time_days": row.get("typical_lead_time_days", "TBD"),
                "product_type": product_type,
            }
    return {
        "supplier_name": fallback_name,
        "supplier_contact": "TBD",
        "order_channel": "TBD",
        "order_cutoff_time": "TBD",
        "typical_lead_time_days": "TBD",
        "product_type": product_type,
    }


def parse_sales_metrics(rows: list[list[str]]) -> dict[str, object]:
    if not rows:
        return {"top": [], "weak": []}
    header_idx = 0
    name_col = -1
    qty_col = -1
    rev_col = -1
    unit_col = -1
    for ridx in range(min(len(rows), 10)):
        header = [c.strip().lower() for c in rows[ridx]]
        for i, h in enumerate(header):
            if name_col == -1 and any(k in h for k in ("наимен", "товар", "номенклат", "пози")):
                name_col = i
            if qty_col == -1 and any(k in h for k in ("кол", "шт", "qty", "кол-во")):
                qty_col = i
            if rev_col == -1 and any(k in h for k in ("выруч", "сумм", "оборот", "sales")):
                rev_col = i
            if unit_col == -1 and any(k in h for k in ("ед.", "ед изм", "unit")):
                unit_col = i
        if name_col != -1 and (qty_col != -1 or rev_col != -1):
            header_idx = ridx
            break
    if name_col == -1 and len(rows) > 12:
        probe = rows[9] if len(rows) > 9 else []
        if len(probe) >= 3 and any("кол" in str(cell).lower() for cell in probe):
            header_idx = 11
            name_col = 0
            qty_col = 2
            unit_col = 3
    if name_col == -1:
        name_col = 0
    items: list[tuple[str, float, float]] = []
    for r in rows[header_idx + 1 :]:
        if len(r) <= name_col:
            continue
        name = (r[name_col] or "").strip()
        if not name:
            continue
        if name.lower().startswith("построен:") or name.lower().startswith("наша компания:"):
            continue
        unit = ""
        if unit_col != -1 and len(r) > unit_col:
            unit = (r[unit_col] or "").strip().lower()
        qty = 0.0
        rev = 0.0
        if qty_col != -1 and len(r) > qty_col:
            qty = parse_number(r[qty_col] or "") or 0.0
        if rev_col != -1 and len(r) > rev_col:
            rev = parse_number(r[rev_col] or "") or 0.0
        if rev <= 0 and len(r) > 5:
            rev = parse_number(r[5] or "") or 0.0
        if unit_col != -1 and unit not in {"шт", "кг", "л", "порц", "порц.", "упак", "упак."}:
            continue
        if qty <= 0 and rev <= 0:
            continue
        items.append((name, qty, rev))
    if not items:
        return {"top": [], "weak": []}
    by_qty = sorted(items, key=lambda x: (x[1], x[2]), reverse=True)
    weak = sorted(items, key=lambda x: (x[1], x[2]))[:5]
    return {"top": by_qty[:5], "weak": weak}


def parse_catalog_prices(rows: list[list[str]]) -> dict[str, float]:
    if not rows:
        return {}
    name_col = -1
    price_col = -1
    header_idx = 0
    for ridx in range(min(len(rows), 10)):
        header = [c.strip().lower() for c in rows[ridx]]
        for i, h in enumerate(header):
            if name_col == -1 and any(k in h for k in ("наимен", "товар", "номенклат", "пози")):
                name_col = i
            if price_col == -1 and any(k in h for k in ("цена", "стоим", "price")):
                price_col = i
        if name_col != -1 and price_col != -1:
            header_idx = ridx
            break
    if name_col == -1 or price_col == -1:
        return {}
    out: dict[str, float] = {}
    for r in rows[header_idx + 1 :]:
        if len(r) <= max(name_col, price_col):
            continue
        name = (r[name_col] or "").strip()
        price = parse_number(r[price_col] or "")
        if not name or price is None or price <= 0:
            continue
        out[normalize_item_name(name)] = float(price)
    return out


def price_delta_from_catalog(insights: list[dict]) -> dict[str, list[str]]:
    catalogs = [i for i in insights if i.get("report_type") == "catalog" and i.get("catalog_prices")]
    if len(catalogs) < 2:
        return {"up": [], "down": []}
    def _mtime(it: dict) -> float:
        rel = str(it.get("source_rel") or "")
        p = ROOT / rel if rel else None
        return p.stat().st_mtime if p and p.exists() else 0.0
    catalogs = sorted(catalogs, key=_mtime)
    prev = catalogs[-2].get("catalog_prices") or {}
    curr = catalogs[-1].get("catalog_prices") or {}
    up: list[tuple[str, float]] = []
    down: list[tuple[str, float]] = []
    for name, new_price in curr.items():
        old_price = prev.get(name)
        if old_price is None or old_price <= 0:
            continue
        delta_pct = ((new_price - old_price) / old_price) * 100.0
        if abs(delta_pct) < 0.5:
            continue
        if delta_pct > 0:
            up.append((name, delta_pct))
        else:
            down.append((name, delta_pct))
    up.sort(key=lambda x: x[1], reverse=True)
    down.sort(key=lambda x: x[1])
    return {
        "up": [f"{n.title()} (+{d:.1f}%)" for n, d in up[:5]],
        "down": [f"{n.title()} ({d:.1f}%)" for n, d in down[:5]],
    }


def price_delta_from_invoices() -> dict[str, list[str]]:
    items = collect_invoice_line_items()
    if not items:
        return {"up": [], "down": []}

    by_sku: dict[tuple[str, str], list[dict[str, object]]] = {}
    for item in items:
        price = float(item.get("price") or 0)
        if price <= 0 or price == 0.01:
            continue
        key = (
            str(item.get("seller") or "").strip(),
            normalize_item_name(str(item.get("name") or "")),
        )
        by_sku.setdefault(key, []).append(item)

    up: list[tuple[str, float]] = []
    down: list[tuple[str, float]] = []
    for (seller, _), rows in by_sku.items():
        rows = [r for r in rows if r.get("invoice_dt") is not None]
        if len(rows) < 2:
            continue
        rows.sort(key=lambda r: r.get("invoice_dt"))
        prev = rows[-2]
        curr = rows[-1]
        old_price = float(prev.get("price") or 0)
        new_price = float(curr.get("price") or 0)
        if old_price <= 0 or new_price <= 0:
            continue
        delta_pct = ((new_price - old_price) / old_price) * 100.0
        if abs(delta_pct) < 0.5:
            continue
        item_name = str(curr.get("name") or "").strip()
        label = f"{seller}: {item_name} ({old_price:.2f} -> {new_price:.2f})"
        if delta_pct > 0:
            up.append((label, delta_pct))
        else:
            down.append((label, delta_pct))

    up.sort(key=lambda x: x[1], reverse=True)
    down.sort(key=lambda x: x[1])
    return {
        "up": [f"{label} (+{delta:.1f}%)" for label, delta in up[:5]],
        "down": [f"{label} ({delta:.1f}%)" for label, delta in down[:5]],
    }


def consumption_alerts(insights: list[dict]) -> list[str]:
    stocks = [i for i in insights if i.get("report_type") in {"stock", "inventory", "beans"} and i.get("stock_items")]
    if len(stocks) < 2:
        return []
    def _mtime(it: dict) -> float:
        rel = str(it.get("source_rel") or "")
        p = ROOT / rel if rel else None
        return p.stat().st_mtime if p and p.exists() else 0.0
    stocks = sorted(stocks, key=_mtime)
    prev = stocks[-2].get("stock_items") or {}
    curr = stocks[-1].get("stock_items") or {}
    alerts: list[tuple[str, int]] = []
    for name, old_qty in prev.items():
        new_qty = curr.get(name)
        if new_qty is None:
            continue
        delta = int(old_qty) - int(new_qty)
        if delta <= 0:
            continue
        if any(k in name for k in ("моющ", "хоз", "чист", "салфет", "перчат")) and delta >= 3:
            alerts.append((name, delta))
    alerts.sort(key=lambda x: x[1], reverse=True)
    return [f"{n.title()}: расход {d} шт за период" for n, d in alerts[:5]]


def target_stock_level(abc: str) -> int:
    return 10


def order_urgency(qty_now: int, abc: str) -> tuple[str, str]:
    abc_norm = (abc or "").strip().upper()
    if qty_now <= 1:
        return "critical", "сегодня до 18:00"
    if abc_norm == "A" and qty_now <= 3:
        return "critical", "сегодня до 18:00"
    if abc_norm in {"A", "B"} and qty_now <= 4:
        return "planned", "в ближайшие 2-3 дня"
    return "planned", "в течение недели"


def format_money(value: float) -> str:
    return f"{value:,.0f} ₽".replace(",", " ")


def executive_summary_lines(
    critical_orders: list[dict[str, object]],
    planned_orders: list[dict[str, object]],
    reduce_or_stop: list[str],
    manual_review: list[str],
    price_delta: dict[str, list[str]],
    consumption: list[str],
    has_abc: bool,
) -> list[str]:
    suppliers = sorted(
        {
            str(item.get("supplier_name", "")).strip()
            for item in [*critical_orders, *planned_orders]
            if str(item.get("supplier_name", "")).strip()
        }
    )
    if critical_orders:
        main_risk = "out-of-stock по части ассортиментных позиций"
    elif manual_review:
        main_risk = "часть решений пока держится на ручной проверке данных"
    elif reduce_or_stop:
        main_risk = "часть денег может застревать в слабых позициях"
    else:
        main_risk = "критичных дефицитов по текущим данным не видно"

    lines = [
        f"Критичных SKU: {len(critical_orders)}; плановых дозаказов: {len(planned_orders)}; поставщиков в цикле: {len(suppliers)}.",
        f"Главный риск периода: {main_risk}.",
    ]
    if suppliers:
        lines.append(f"Ключевые поставщики цикла: {', '.join(suppliers)}.")
    if reduce_or_stop:
        lines.append(f"Позиции на пересмотр закупки/ассортимента: {len(reduce_or_stop)}.")
    if price_delta.get("up"):
        lines.append(f"Есть сигнал роста цен по {len(price_delta.get('up', []))} позициям.")
    if consumption:
        lines.append("Есть сигнал по повышенному расходу хозтоваров/расходников.")
    if not has_abc:
        lines.append("ABC-контур не загружен: приоритеты закупки пока временно опираются на остатки и продажи.")
    return lines


def build_order_bullets(orders: list[dict[str, object]]) -> list[str]:
    bullets: list[str] = []
    for item in orders:
        bullets.append(
            f"{item['name']}: сейчас {item['qty_now']} шт, заказать {item['qty_to_order']} шт, "
            f"дедлайн {item['deadline']}, риск: {item['risk']}"
        )
    return bullets


def bucket_orders_by_supplier(orders: list[dict[str, object]]) -> list[tuple[str, list[dict[str, object]]]]:
    buckets: dict[str, list[dict[str, object]]] = {}
    for item in orders:
        supplier_name = str(item.get("supplier_name", "")).strip() or "Уточнить у Жанны"
        buckets.setdefault(supplier_name, []).append(item)
    return sorted(buckets.items(), key=lambda kv: kv[0].lower())


def build_supplier_order_message(supplier_name: str, items: list[dict[str, object]]) -> str:
    contact = str(items[0].get("supplier_contact", "")).strip() or "TBD"
    lines = [
        f"Здравствуйте. Хотим оформить заказ по поставщику {supplier_name}:",
        "",
    ]
    for idx, item in enumerate(items, start=1):
        lines.append(f"{idx}. {item['name']} — {item['qty_to_order']} шт.")
    lines.extend(
        [
            "",
            "Просьба подтвердить наличие, цену и ближайшую дату отгрузки.",
        ]
    )
    if contact and contact != "TBD":
        lines.extend(["", f"Контакт в контуре: {contact}"])
    return "\n".join(lines)


def format_supplier_channel(channel: str, contact: str) -> str:
    channel = (channel or "").strip()
    contact = (contact or "").strip()
    if channel in {"", "TBD"} and contact in {"", "TBD"}:
        return "TBD"
    if channel in {"", "TBD"}:
        return contact
    if contact in {"", "TBD"}:
        return channel
    return f"{channel} {contact}"


def build_smart_analytics(insights: list[dict]) -> dict:
    """Кросс-анализ остатков + ABC категорий."""
    # Собираем все остатки (позиция -> остаток)
    stock_map: dict[str, int] = {}
    stock_map_raw: dict[str, str] = {}
    for item in insights:
        if item.get("report_type") in ("stock", "inventory", "beans"):
            for name, qty in (item.get("top_low_items") or []):
                key = normalize_item_name(name)
                if key not in stock_map or stock_map[key] > qty:
                    stock_map[key] = qty
                    stock_map_raw[key] = str(name).strip()

    # Собираем ABC категории
    abc_map: dict[str, str] = {}
    for item in insights:
        if item.get("report_type") == "abc":
            raw = item.get("abc_categories") or {}
            for k, v in raw.items():
                abc_map[normalize_item_name(str(k))] = str(v).strip().upper()

    if not stock_map:
        return {}

    urgent: list[str] = []
    attention: list[str] = []
    excess: list[str] = []
    no_abc: list[str] = []

    for name_low, qty in sorted(stock_map.items(), key=lambda x: x[1]):
        display_name = canonical_item_label(stock_map_raw.get(name_low) or name_low.title())
        cat, _ = resolve_abc_category(display_name, abc_map)
        display = f"{display_name}: {qty} шт"
        if cat == "A":
            urgent.append(display)
        elif cat == "B":
            attention.append(display)
        elif cat == "C":
            excess.append(display)
        else:
            no_abc.append(display)

    # Позиции C категории с нормальными остатками
    for name_low, cat in abc_map.items():
        if cat == "C" and name_low not in stock_map:
            excess.append(f"{name_low.title()} (C-категория, в реестре)")

    sales_top_names: set[str] = set()
    sales_weak_names: set[str] = set()
    for item in insights:
        if item.get("report_type") != "sales":
            continue
        sm = item.get("sales_metrics") or {}
        for name, _, _ in sm.get("top", [])[:10]:
            sales_top_names.add(normalize_item_name(str(name)))
        for name, _, _ in sm.get("weak", [])[:10]:
            sales_weak_names.add(normalize_item_name(str(name)))

    order_now: list[str] = []
    order_by_supplier: dict[str, dict[str, object]] = {}
    critical_orders: list[dict[str, object]] = []
    planned_orders: list[dict[str, object]] = []
    manual_review: list[str] = []
    reduce_or_stop: list[str] = []
    passive_monitoring: list[str] = []

    for name_low, qty in sorted(stock_map.items(), key=lambda x: x[1]):
        raw_name = canonical_item_label(stock_map_raw.get(name_low) or name_low.title())
        cat, cat_mode = resolve_abc_category(raw_name, abc_map)
        if not cat:
            manual_review.append(
                f"{raw_name}: {qty} шт: нет ABC-категории, решение пока по остатку."
            )
            cat = "A"
        elif cat_mode == "family":
            manual_review.append(
                f"{raw_name}: ABC-категория {cat} подтянута по семейству SKU, проверить совпадение с weekly ABC."
            )
        min_target = target_stock_level(cat)
        to_order = max(0, min_target - int(qty))
        if to_order <= 0:
            continue
        supplier = infer_supplier_for_item(raw_name)
        supplier_name = supplier.get("supplier_name", "Уточнить у Жанны")
        contact = supplier.get("supplier_contact", "TBD")
        channel = supplier.get("order_channel", "TBD")
        cutoff = supplier.get("order_cutoff_time", "до 18:00")
        lead_time = supplier.get("typical_lead_time_days", "TBD")
        product_type = supplier.get("product_type", "другое")
        urgency, deadline = order_urgency(int(qty), cat)
        risk = "стоп продаж по позиции" if int(qty) <= 1 else "риск дефицита в ближайшие дни"
        reason = f"остаток {qty} шт при целевом уровне {min_target} шт, ABC {cat}, lead time {lead_time} дн."

        if product_type in {"десерты", "кухня"}:
            passive_monitoring.append(
                f"{raw_name}: не выводить в заказ, отслеживать цены поставки и продажи"
            )
            continue

        order_now.append(
            f"{raw_name}: заказать {to_order} шт · {supplier_name} · {channel} {contact} · {cutoff}"
        )
        order_row = {
            "name": raw_name,
            "qty_now": int(qty),
            "qty_to_order": int(to_order),
            "abc": cat,
            "supplier_name": supplier_name,
            "supplier_contact": contact,
            "order_channel": channel,
            "order_cutoff_time": cutoff,
            "typical_lead_time_days": lead_time,
            "deadline": deadline,
            "risk": risk,
            "reason": reason,
            "product_type": product_type,
        }
        if contact == "TBD" or channel == "TBD":
            manual_review.append(
                f"{raw_name}: нет подтверждённого контакта/канала поставщика ({supplier_name})"
            )
        if urgency == "critical":
            critical_orders.append(order_row)
        else:
            planned_orders.append(order_row)
        bucket = order_by_supplier.setdefault(
            supplier_name,
            {
                "supplier_name": supplier_name,
                "supplier_contact": contact,
                "order_channel": channel,
                "order_cutoff_time": cutoff,
                "items": [],
            },
        )
        bucket_items = bucket.get("items")
        if isinstance(bucket_items, list):
            bucket_items.append(
                {
                    "name": raw_name,
                    "qty_now": int(qty),
                    "qty_to_order": int(to_order),
                    "abc": cat,
                    "product_type": product_type,
                    "deadline": deadline,
                }
            )
    order_now = order_now[:8]

    for name_low in sorted(abc_map.keys()):
        if abc_map.get(name_low) != "C":
            continue
        if name_low in sales_top_names:
            continue
        display_name = canonical_item_label(stock_map_raw.get(name_low) or name_low.title())
        if name_low in sales_weak_names:
            reduce_or_stop.append(
                f"{display_name}: слабая продажа + C-категория, не расширять закупку до следующего цикла"
            )
        elif name_low not in stock_map:
            reduce_or_stop.append(
                f"{display_name}: C-категория, держать минимальный остаток без доп. закупки"
            )

    for weak_name in sorted(sales_weak_names):
        product_type = classify_item_product_type(weak_name)
        if product_type not in {"десерты", "кухня"}:
            continue
        display_name = canonical_item_label(stock_map_raw.get(weak_name) or weak_name.title())
        reduce_or_stop.append(
            f"{display_name}: слабая продажа в десертном/кухонном контуре, проверить сокращение или снятие"
        )

    critical_orders = sorted(
        critical_orders,
        key=lambda x: (int(x.get("qty_now", 0)), str(x.get("name", ""))),
    )
    planned_orders = sorted(
        planned_orders,
        key=lambda x: (int(x.get("qty_now", 0)), str(x.get("name", ""))),
    )
    manual_review = list(dict.fromkeys(manual_review))[:8]
    reduce_or_stop = list(dict.fromkeys(reduce_or_stop))[:8]
    passive_monitoring = list(dict.fromkeys(passive_monitoring))[:8]

    return {
        "urgent": urgent[:8],
        "attention": attention[:6],
        "excess": excess[:5],
        "no_abc": no_abc[:5],
        "has_abc": bool(abc_map),
        "order_now": order_now,
        "order_by_supplier": list(order_by_supplier.values())[:8],
        "critical_orders": critical_orders,
        "planned_orders": planned_orders,
        "manual_review": manual_review,
        "reduce_or_stop": reduce_or_stop,
        "passive_monitoring": passive_monitoring,
        "stock_total_items": len(stock_map),
        "abc_total_items": len(abc_map),
    }


def detect_data_gaps(insights: list[dict]) -> list[str]:
    """Каких данных не хватает для управленческого отчёта руководителя."""
    types = {str(item.get("report_type", "")) for item in insights}
    gaps: list[str] = []

    if "invoice" not in types:
        gaps.append("Нет накладных в структурированном виде: не считаем точную закупочную себестоимость по SKU.")
    else:
        gaps.append("Накладные PDF уже частично разобраны до SKU-уровня, но extraction coverage и качество price-ledger ещё нужно усиливать.")

    if "sales" not in types:
        gaps.append("Нет отчёта продаж по SKU за период: нельзя считать спрос и оборачиваемость по позициям.")

    if "revenue" not in types:
        gaps.append("Нет детальной выручки по точкам/дням: ограничен контроль динамики точек.")

    gaps.append("Нет отчёта списаний/брака: не видно потери и причины перерасхода.")
    gaps.append("Нет факта поставок (lead time, недопоставки): нельзя оценить надёжность поставщиков.")
    return gaps[:5]


def make_card(source: Path) -> tuple[Path, Path, dict]:
    CARDS_DIR.mkdir(parents=True, exist_ok=True)
    BOT_REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    rel = source.relative_to(ROOT)
    stamp = datetime.fromtimestamp(source.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
    source_sig = hashlib.sha1(rel.as_posix().encode("utf-8")).hexdigest()[:8]
    slug = f"{sanitize_slug(source.stem)}-{source_sig}"
    card_path = CARDS_DIR / f"WH.CARD.{slug}.md"
    bot_path = BOT_REPORTS_DIR / f"WH.BOT.{slug}.md"

    suffix = source.suffix.lower()
    body_lines = [
        "---",
        "type: warehouse-report-card",
        f"source: {rel.as_posix()}",
        f"updated: {stamp}",
        "status: active",
        "---",
        "",
        f"# Карточка отчета: {source.name}",
        "",
        f"- Источник: `{rel.as_posix()}`",
        f"- Обновлен: `{stamp}`",
        "",
    ]

    insight: dict[str, object] = {
        "title": source.name,
        "report_type": detect_report_type(source.name),
        "source_rel": rel.as_posix(),
        "card_rel": "",
        "bot_card_rel": "",
    }

    report_type = detect_report_type(source.name)
    if (suffix == ".pdf") and (report_type == "invoice"):
        payload = extract_pdf_payload(source)
        text = str(payload.get("text", "") or "")
        metrics = parse_invoice_pdf_metrics(text) if text else {}
        period = extract_period_from_text(source.name)

        insight["report_type"] = "invoice"
        insight["rows"] = int(metrics.get("line_count", 0) or 0)
        insight["period"] = period
        insight["invoice_numbers"] = metrics.get("invoice_numbers", [])
        insight["invoice_money_top"] = metrics.get("money_top", [])

        body_lines.extend(
            [
                "## Накладные (PDF): авторазбор",
                f"- Период: **{period}**",
                f"- Метод извлечения: **{payload.get('method') or 'n/a'}**",
                f"- Страниц: **{payload.get('pages') or 0}**",
                f"- Извлечено строк: **{metrics.get('line_count', 0) if metrics else 0}**",
                f"- Извлечено символов: **{metrics.get('char_count', 0) if metrics else 0}**",
            ]
        )
        if metrics:
            nums = metrics.get("invoice_numbers") or []
            if nums:
                body_lines.append(f"- Номера накладных: **{', '.join(str(x) for x in nums[:10])}**")
            dts = metrics.get("dates") or []
            if dts:
                body_lines.append(f"- Обнаруженные даты: **{', '.join(str(x) for x in dts[:8])}**")
            vendors = metrics.get("vendors") or []
            if vendors:
                body_lines.append("- Поставщики (по тексту):")
                for v in vendors:
                    body_lines.append(f"  - {v}")
            mtop = metrics.get("money_top") or []
            if mtop:
                body_lines.append("- Крупные суммы (по тексту, авто):")
                for m in mtop[:6]:
                    body_lines.append(f"  - {m:,.2f}".replace(",", " "))
            preview = metrics.get("preview_lines") or []
            if preview:
                body_lines.append("")
                body_lines.append("## Превью извлечённого текста")
                for ln in preview[:10]:
                    body_lines.append(f"- {ln}")
        if not text:
            body_lines.append("- Не удалось извлечь текст из PDF автоматически.")
            if payload.get("error"):
                body_lines.append(f"- Ошибка: `{payload.get('error')}`")
    else:
        rows = parse_tabular_rows(source)
        if ("Комментарий" in source.name) or ("Комментари" in source.name):
            bullets = parse_comment_bullets(rows)
            insight["comment_count"] = len(bullets)
            insight["comment_preview"] = bullets[:3]
            body_lines.append("## Выжимка комментариев")
            if bullets:
                for b in bullets:
                    body_lines.append(f"- {b}")
            else:
                body_lines.append("- Нет содержимого для выжимки.")
        else:
            if report_type == "abc":
                abc_payload: dict[str, object]
                if suffix == ".xlsx":
                    sheets = parse_xlsx_sheets(source)
                    parsed = [
                        parse_abc_payload(
                            list(sheet.get("rows") or []),
                            sheet_name=str(sheet.get("title") or ""),
                        )
                        for sheet in sheets
                    ]
                    abc_payload = max(
                        parsed,
                        key=lambda item: (
                            int(item.get("matched_rows", 0) or 0),
                            -int(item.get("unmatched_rows", 0) or 0),
                            int(item.get("total_rows", 0) or 0),
                        ),
                    )
                else:
                    abc_payload = parse_abc_payload(rows)

                categories = dict(abc_payload.get("categories") or {})
                insight["abc_categories"] = categories
                insight["abc_sheet_name"] = str(abc_payload.get("sheet_name") or "")
                insight["abc_matched_rows"] = int(abc_payload.get("matched_rows", 0) or 0)
                insight["abc_unmatched_rows"] = int(abc_payload.get("unmatched_rows", 0) or 0)
                insight["abc_status"] = str(abc_payload.get("status") or "")
                insight["abc_unmatched_preview"] = list(abc_payload.get("unmatched_preview") or [])
                counts = {"A": 0, "B": 0, "C": 0}
                for cat in categories.values():
                    if cat in counts:
                        counts[cat] += 1
                body_lines.extend([
                    "## ABC Анализ",
                    f"- Статус разбора: **{insight['abc_status'] or 'n/a'}**",
                    (
                        f"- Лист: **{insight['abc_sheet_name']}**"
                        if insight.get("abc_sheet_name")
                        else "- Лист: **n/a**"
                    ),
                    f"- Matched SKU: **{insight['abc_matched_rows']}**",
                    f"- Unmatched строк: **{insight['abc_unmatched_rows']}**",
                    f"- Позиций A (приоритет): **{counts['A']}**",
                    f"- Позиций B (средние): **{counts['B']}**",
                    f"- Позиций C (низкий приоритет): **{counts['C']}**",
                    "",
                    "## Топ категории A",
                ])
                a_items = [name for name, cat in categories.items() if cat == "A"][:10]
                for name in a_items:
                    body_lines.append(f"- {name.title()}")
                if not a_items:
                    body_lines.append("- Не найдено.")
                unmatched_preview = list(insight.get("abc_unmatched_preview") or [])
                if unmatched_preview:
                    body_lines.extend(["", "## Требуют ручной сверки"])
                    for name in unmatched_preview[:6]:
                        body_lines.append(f"- {name}")
            elif report_type in {"stock", "inventory", "beans"}:
                metrics = parse_stock_metrics(rows)
                if metrics["rows"] > 0:
                    insight["rows"] = metrics["rows"]
                    insight["low_stock_count"] = metrics["low_stock_count"]
                    insight["top_low_items"] = metrics["low_stock_items"]
                    insight["stock_items"] = metrics.get("item_totals") or {}
                    body_lines.extend(
                        [
                            "## Метрики",
                            f"- Учетных позиций: **{metrics['rows']}**",
                            f"- Низкие остатки (<=3): **{metrics['low_stock_count']}**",
                            (
                                "- Сумма по точкам: "
                                f"Тургенева={metrics['totals_by_location']['Тургенева']}, "
                                f"Луговая={metrics['totals_by_location']['Луговая']}, "
                                f"Самокиша={metrics['totals_by_location']['Самокиша']}"
                            ),
                            "",
                            "## Низкий остаток (top 12)",
                        ]
                    )
                    low_items = metrics["low_stock_items"][:12]
                    if low_items:
                        for name, total in low_items:
                            body_lines.append(f"- {name}: {total}")
                    else:
                        body_lines.append("- Низких остатков не найдено.")
                else:
                    # Файл помечен как инвентаризация, но не является SKU-остатками.
                    # Делаем понятную бизнес-выжимку вместо пустых нулей.
                    profile = parse_table_profile(rows)
                    extra = parse_non_stock_inventory_metrics(rows)
                    insight["report_type"] = "inventory_non_stock"
                    insight["rows"] = int(profile["rows_data"])
                    insight["low_stock_count"] = 0
                    insight["top_low_items"] = []
                    insight["inventory_people_count"] = int(extra.get("people_count", 0) or 0)
                    insight["inventory_money_cells_count"] = int(extra.get("money_cells_count", 0) or 0)
                    insight["inventory_debt_notes"] = extra.get("debt_notes", [])

                    body_lines.extend(
                        [
                            "## Тип файла",
                            "- Это не карточка товарных остатков по SKU.",
                            "- Это сводная инвентаризационно-финансовая таблица (по сотрудникам/периодам).",
                            "",
                            "## Выжимка по данным",
                            f"- Строк данных: **{profile['rows_data']}**",
                            f"- Столбцов: **{profile['cols']}**",
                            f"- Сотрудников в таблице (по автоопределению): **{extra['people_count']}**",
                            f"- Денежных значений: **{extra['money_cells_count']}**",
                            f"- Сумма денежных значений (справочно): **{float(extra['money_cells_sum']):,.2f} ₽**".replace(",", " "),
                        ]
                    )
                    people = extra.get("people_top") or []
                    if people:
                        body_lines.append(f"- Имена (пример): **{', '.join(str(x) for x in people[:8])}**")
                    debts = extra.get("debt_notes") or []
                    if debts:
                        body_lines.append("")
                        body_lines.append("## Пометки по долгам/корректировкам")
                        for d in debts:
                            body_lines.append(f"- {d}")
            else:
                profile = parse_table_profile(rows)
                period = extract_period_from_text(source.name)
                insight["rows"] = int(profile["rows_data"])
                insight["low_stock_count"] = 0
                insight["top_low_items"] = []
                insight["period"] = period
                insight["table_headers"] = profile["headers"]
                insight["top_numeric"] = profile["top_numeric"]
                if report_type == "sales":
                    insight["sales_metrics"] = parse_sales_metrics(rows)
                if report_type == "catalog":
                    insight["catalog_prices"] = parse_catalog_prices(rows)

                type_human = {
                    "sales": "Продажи",
                    "revenue": "Выручка",
                    "catalog": "Каталог",
                    "invoice": "Накладные",
                    "other": "Табличный отчёт",
                }.get(report_type, "Табличный отчёт")

                body_lines.extend(
                    [
                        f"## {type_human}: структурная выжимка",
                        f"- Период: **{period}**",
                        f"- Строк данных: **{profile['rows_data']}**",
                        f"- Столбцов: **{profile['cols']}**",
                    ]
                )
                headers = profile.get("headers") or []
                if headers:
                    body_lines.append(f"- Ключевые колонки: **{', '.join(str(h) for h in headers)}**")
                top_numeric = profile.get("top_numeric") or []
                if top_numeric:
                    body_lines.append("")
                    body_lines.append("## Числовые итоги (автооценка)")
                    for _, label, total in top_numeric:
                        body_lines.append(f"- {label}: **{total:,.2f}**".replace(",", " "))
                if report_type == "invoice":
                    body_lines.append("")
                    body_lines.append("## Важно")
                    body_lines.append("- Для PDF-накладных нужна OCR-нормализация для SKU-level аналитики (в этом цикле только базовая выжимка).")

    body_lines.extend(
        [
            "",
            "## Следующий шаг",
            "- Проверить потребности точки по заявке и сверить с минимальными остатками.",
            "- При необходимости добавить SKU в ближайшую закупку.",
            "",
        ]
    )

    text = "\n".join(body_lines)
    card_path.write_text(text, encoding="utf-8")
    bot_path.write_text(text, encoding="utf-8")
    insight["card_rel"] = card_path.relative_to(ROOT).as_posix()
    insight["bot_card_rel"] = bot_path.relative_to(ROOT).as_posix()
    return card_path, bot_path, insight


def build_latest_summary(
    cards: Iterable[Path],
    bot_cards: Iterable[Path],
    hours: int,
    run_stats: dict[str, int],
    insights: list[dict],
) -> str:
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    analytics = build_smart_analytics(insights)
    price_delta = price_delta_from_invoices()
    if (not price_delta.get("up")) and (not price_delta.get("down")):
        price_delta = price_delta_from_catalog(insights)
    consumption = consumption_alerts(insights)
    gaps = detect_data_gaps(insights)
    sales_top: list[str] = []
    sales_weak: list[str] = []
    for item in insights:
        if item.get("report_type") != "sales":
            continue
        sm = item.get("sales_metrics") or {}
        for name, qty, rev in sm.get("top", [])[:5]:
            sales_top.append(f"{name}: {qty:.0f} шт / {format_money(float(rev))}")
        for name, qty, rev in sm.get("weak", [])[:5]:
            sales_weak.append(f"{name}: {qty:.0f} шт / {format_money(float(rev))}")
    sales_top = list(dict.fromkeys(sales_top))[:5]
    sales_weak = list(dict.fromkeys(sales_weak))[:5]
    critical_orders = analytics.get("critical_orders") or []
    planned_orders = analytics.get("planned_orders") or []
    reduce_or_stop = analytics.get("reduce_or_stop") or []
    passive_monitoring = analytics.get("passive_monitoring") or []
    manual_review = list(analytics.get("manual_review") or [])
    if analytics.get("no_abc"):
        manual_review.extend(
            [f"{item}: нет ABC-категории, решение пока по остатку." for item in analytics.get("no_abc", [])[:5]]
        )
    manual_review = list(dict.fromkeys(manual_review))[:8]
    executive = executive_summary_lines(
        critical_orders=critical_orders,
        planned_orders=planned_orders,
        reduce_or_stop=reduce_or_stop,
        manual_review=manual_review,
        price_delta=price_delta,
        consumption=consumption,
        has_abc=bool(analytics.get("has_abc")),
    )

    lines = [
        "---",
        "type: warehouse-manager-report",
        f"updated: {now}",
        "---",
        "",
        "# Warehouse Manager Report",
        "",
        f"- Окно анализа: последние `{hours}` ч",
        f"- Источников в окне: **{run_stats['received']}**",
        f"- Обработано: **{run_stats['processed']}**",
        f"- Ошибок парсинга: **{run_stats['error']}**",
        f"- SKU с остатками в анализе: **{analytics.get('stock_total_items', 0)}**",
        f"- SKU с ABC-категорией: **{analytics.get('abc_total_items', 0)}**",
        "",
        "## Executive Summary",
    ]
    for item in executive:
        lines.append(f"- {item}")

    lines.extend(["", "## Срочно заказать"])
    if critical_orders:
        for supplier_name, items in bucket_orders_by_supplier(critical_orders):
            channel = str(items[0].get("order_channel", "")).strip()
            contact = str(items[0].get("supplier_contact", "")).strip()
            channel_text = format_supplier_channel(channel, contact)
            common_deadline = str(items[0].get("deadline", "")).strip() or "уточнить"
            lines.extend(
                [
                    f"- **{supplier_name}**",
                    f"  Канал заказа: **{channel_text}**.",
                    f"  Базовый дедлайн: **{common_deadline}**.",
                ]
            )
            for item in items:
                item_deadline = str(item.get("deadline", "")).strip() or common_deadline
                deadline_note = ""
                if item_deadline != common_deadline:
                    deadline_note = f", дедлайн: **{item_deadline}**"
                lines.append(
                    f"  - {item['name']}: сейчас **{item['qty_now']} шт**, заказать **{item['qty_to_order']} шт**, риск: **{item['risk']}**{deadline_note}."
                )
    else:
        lines.append("- Критичных позиций для заказа сейчас нет.")

    lines.extend(["", "## Готовые заявки поставщикам"])
    if critical_orders:
        for supplier_name, items in bucket_orders_by_supplier(critical_orders):
            lines.append(f"- **{supplier_name}**")
            for msg_line in build_supplier_order_message(supplier_name, items).splitlines():
                if msg_line:
                    lines.append(f"  {msg_line}")
                else:
                    lines.append("  ")
    else:
        lines.append("- Готовых срочных заявок поставщикам в этом цикле нет.")

    lines.extend(["", "## Планово заказать"])
    if planned_orders:
        for supplier_name, items in bucket_orders_by_supplier(planned_orders):
            common_deadline = str(items[0].get("deadline", "")).strip() or "уточнить"
            lines.extend(
                [
                    f"- **{supplier_name}**",
                    f"  Базовый дедлайн: **{common_deadline}**.",
                ]
            )
            for item in items:
                item_deadline = str(item.get("deadline", "")).strip() or common_deadline
                deadline_note = ""
                if item_deadline != common_deadline:
                    deadline_note = f" Дедлайн по позиции: **{item_deadline}**."
                lines.append(
                    f"  - {item['name']}: сейчас **{item['qty_now']} шт**, целевой уровень **{target_stock_level(str(item['abc']))} шт**, дозаказать **{item['qty_to_order']} шт**."
                )
                lines.append(f"    Причина: {item['reason']}.{deadline_note}")
    else:
        lines.append("- Плановых позиций к дозаказу не выявлено.")

    lines.extend(["", "## Не заказывать / сократить"])
    if reduce_or_stop:
        for item in reduce_or_stop:
            lines.append(f"- {item}")
    else:
        lines.append("- Явных позиций под сокращение закупки пока не выявлено.")

    lines.extend(["", "## Десерты и кухня: только аналитика"])
    if passive_monitoring:
        for item in passive_monitoring:
            lines.append(f"- {item}")
    else:
        lines.append("- Десертный и кухонный контур не требует отдельной аналитической пометки в этом цикле.")

    lines.extend(["", "## Продажи и ассортимент"])
    if sales_top:
        lines.append("- Топ продаж:")
        for item in sales_top:
            lines.append(f"  - {item}")
    if sales_weak:
        lines.append("- Слабые позиции:")
        for item in sales_weak:
            lines.append(f"  - {item}")
    if (not sales_top) and (not sales_weak):
        lines.append("- По текущим файлам не удалось уверенно извлечь топ и слабые позиции продаж.")

    lines.extend(["", "## Изменение цен"])
    if price_delta.get("up"):
        lines.append("- Подорожали:")
        for item in price_delta.get("up", [])[:5]:
            lines.append(f"  - {item}")
    if price_delta.get("down"):
        lines.append("- Подешевели:")
        for item in price_delta.get("down", [])[:5]:
            lines.append(f"  - {item}")
    if (not price_delta.get("up")) and (not price_delta.get("down")):
        lines.append("- Недостаточно сопоставимых каталогов для оценки динамики цен.")

    lines.extend(["", "## Операционные сигналы"])
    if consumption:
        for item in consumption:
            lines.append(f"- {item}")
    else:
        lines.append("- Аномальный расход по хозтоварам и моющим не выявлен либо не хватает срезов.")

    lines.extend(["", "## Проверить вручную"])
    if manual_review:
        for item in manual_review:
            lines.append(f"- {item}")
    else:
        lines.append("- Критичных ручных проверок не требуется.")

    lines.extend(["", "## Каких данных не хватает"])
    for item in gaps[:5]:
        lines.append(f"- {item}")

    lines.extend(["", "## Источники и артефакты"])
    lines.append(f"- Реестр: `{REGISTRY_FILE.relative_to(ROOT).as_posix()}`")
    lines.append(f"- Quarantine report: `{DLQ_REPORT.relative_to(ROOT).as_posix()}`")
    if PROCUREMENT_REPORT_FILE.exists():
        lines.append(f"- Закупочный контур: `{PROCUREMENT_REPORT_FILE.relative_to(ROOT).as_posix()}`")
    lines.append("")
    text = "\n".join(lines)
    LATEST_REPORT.write_text(text, encoding="utf-8")
    return text


def build_decision_queue(insights: list[dict], run_stats: dict[str, int], manual_run: bool) -> str:
    """Очередь управленческих решений без документного шума."""
    ts = now_stamp()
    mode = "manual" if manual_run else "auto"
    analytics = build_smart_analytics(insights)
    sales_weak: list[str] = []
    for item in insights:
        if item.get("report_type") != "sales":
            continue
        sm = item.get("sales_metrics") or {}
        for name, qty, rev in sm.get("weak", [])[:6]:
            sales_weak.append(f"{canonical_item_label(str(name))}: {qty:.0f} шт / {format_money(float(rev))}")
    sales_weak = list(dict.fromkeys(sales_weak))[:6]
    gaps = detect_data_gaps(insights)
    lines = [
        "---",
        "type: warehouse-decision-queue",
        f"updated: {ts}",
        f"mode: {mode}",
        "---",
        "",
        "# Warehouse Decision Queue",
        "",
        "## Сводка цикла",
        f"- Входящих: **{run_stats['received']}**",
        f"- Обработано: **{run_stats['processed']}**",
        f"- Дубликатов: **{run_stats['duplicate']}**",
        f"- Ошибок: **{run_stats['error']}**",
        "",
        "## WH.SESSION.001 — Что решить сейчас",
    ]

    critical_orders = analytics.get("critical_orders") or []
    if critical_orders:
        lines.append("- Действие: оформить заказ сегодня.")
        for supplier_name, items in bucket_orders_by_supplier(critical_orders[:8]):
            channel = str(items[0].get("order_channel", "")).strip()
            contact = str(items[0].get("supplier_contact", "")).strip()
            channel_text = format_supplier_channel(channel, contact)
            deadline = str(items[0].get("deadline", "")).strip() or "TBD"
            lines.append(f"  - {supplier_name}")
            lines.append(f"    - канал: {channel_text}")
            lines.append(f"    - базовый дедлайн: {deadline}")
            for item in items:
                item_deadline = str(item.get("deadline", "")).strip() or deadline
                extra_deadline = ""
                if item_deadline != deadline:
                    extra_deadline = f", дедлайн {item_deadline}"
                lines.append(
                    f"    - {item['name']}: сейчас {item['qty_now']} шт, заказать {item['qty_to_order']} шт, риск: {item['risk']}{extra_deadline}"
                )
    else:
        lines.append("- Срочных позиций к заказу не выявлено.")

    lines.extend(["", "## WH.SESSION.001A — Готовая заявка поставщику"])
    if critical_orders:
        for supplier_name, items in bucket_orders_by_supplier(critical_orders[:8]):
            lines.append(f"- {supplier_name}")
            for msg_line in build_supplier_order_message(supplier_name, items).splitlines():
                if msg_line:
                    lines.append(f"  - {msg_line}")
                else:
                    lines.append("  -")
    else:
        lines.append("- Готовых срочных заявок в этом цикле нет.")

    lines.extend(["", "## WH.SESSION.002 — Что проверить вручную"])
    manual_review = list(analytics.get("manual_review") or [])
    if analytics.get("no_abc"):
        manual_review.extend(
            [f"{item}: нет ABC-категории, решение пока по остатку." for item in analytics.get("no_abc", [])[:5]]
        )
    manual_review = list(dict.fromkeys(manual_review))[:8]
    if manual_review:
        for item in manual_review[:8]:
            lines.append(f"- {item}")
    else:
        lines.append("- Критичных ручных проверок нет.")

    lines.extend(["", "## WH.SESSION.003 — Ассортимент и продажи"])
    if sales_weak:
        lines.append("- Слабые позиции для решения оставить / сокращать / снимать:")
        for item in sales_weak:
            lines.append(f"  - {item}")
    else:
        lines.append("- Слабых позиций по текущему набору продаж не выявлено.")

    passive_monitoring = analytics.get("passive_monitoring") or []
    if passive_monitoring:
        lines.append("- Десерты и кухня: только аналитическое наблюдение:")
        for item in passive_monitoring[:8]:
            lines.append(f"  - {item}")

    lines.extend(["", "## WH.SESSION.004 — Каких данных не хватает"])
    for item in gaps[:5]:
        lines.append(f"- {item}")

    text = "\n".join(lines) + "\n"
    DECISION_QUEUE.write_text(text, encoding="utf-8")
    return text


def replay_insights_from_registry(registry: dict[str, dict[str, str]], limit: int = 10) -> list[dict]:
    """Собрать demo-выжимку из уже обработанных карточек, если новых входящих нет."""
    rows = []
    for row in registry.values():
        if (row.get("status") or "").strip() not in {"processed", "duplicate"}:
            continue
        card_rel = (row.get("card_path") or "").strip()
        if not card_rel:
            continue
        try:
            ts = datetime.strptime((row.get("last_processed_at") or "").strip(), "%Y-%m-%d %H:%M:%S")
        except Exception:
            ts = datetime.min
        rows.append((ts, row))
    rows.sort(key=lambda x: x[0], reverse=True)

    insights: list[dict] = []
    for _, row in rows[: max(1, limit)]:
        source_file = (row.get("source_file") or "").strip()
        title = Path(source_file).name if source_file else Path(row.get("card_path") or "report").name
        insights.append(
            {
                "title": title,
                "report_type": (row.get("report_type") or "other").strip(),
                "source_rel": source_file,
                "card_rel": (row.get("card_path") or "").strip(),
                "bot_card_rel": (row.get("bot_card_path") or "").strip(),
            }
        )
    return insights


def rebuild_insights_from_sources(sources: list[Path], limit: int = 20) -> tuple[list[Path], list[Path], list[dict]]:
    cards: list[Path] = []
    bot_cards: list[Path] = []
    insights: list[dict] = []
    for src in sources[: max(1, limit)]:
        try:
            card, bot_card, insight = make_card(src)
        except Exception:
            continue
        cards.append(card)
        bot_cards.append(bot_card)
        insights.append(insight)
    return cards, bot_cards, insights


def append_dlq_entry(source: Path, record_key: str, reason: str) -> str:
    DLQ_DIR.mkdir(parents=True, exist_ok=True)
    DLQ_REPORT.parent.mkdir(parents=True, exist_ok=True)

    stamp = now_stamp()
    rel = source.relative_to(ROOT).as_posix()
    slug = hashlib.sha1(record_key.encode("utf-8")).hexdigest()[:10]
    copy_name = f"{source.stem}-{slug}{source.suffix}"
    dlq_copy = DLQ_DIR / copy_name
    shutil.copy2(source, dlq_copy)

    if not DLQ_REPORT.exists():
        DLQ_REPORT.write_text(
            "\n".join(
                [
                    "---",
                    "type: warehouse-dlq-report",
                    f"updated: {stamp}",
                    "---",
                    "",
                    "# Warehouse DLQ Report",
                    "",
                ]
            )
            + "\n",
            encoding="utf-8",
        )

    existing = DLQ_REPORT.read_text(encoding="utf-8")
    updated = re.sub(
        r"^updated: .*$",
        f"updated: {stamp}",
        existing,
        count=1,
        flags=re.MULTILINE,
    )
    entry = "\n".join(
        [
            f"## {stamp} — {source.name}",
            f"- Source: `{rel}`",
            f"- DLQ copy: `{dlq_copy.relative_to(ROOT).as_posix()}`",
            f"- Reason: `{reason}`",
            "",
        ]
    )
    DLQ_REPORT.write_text(updated + entry, encoding="utf-8")
    return dlq_copy.relative_to(ROOT).as_posix()


def read_env_from_telegram_bot() -> dict[str, str]:
    env: dict[str, str] = {}
    env_path = ROOT / "telegram-bot" / ".env"
    if not env_path.exists():
        return env
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def read_key_value_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.exists():
        return env
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def read_first_line(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8", errors="ignore").strip()


def resolve_telegram_credentials() -> tuple[str, str]:
    env_file = read_env_from_telegram_bot()
    aist_env = read_key_value_env(Path.home() / ".config" / "aist" / "env")
    exocortex_chat = read_first_line(Path.home() / ".config" / "exocortex" / "telegram-chat-id")
    exocortex_token = read_first_line(Path.home() / ".config" / "exocortex" / "telegram-token")
    # 1) Явные env (без смешивания со сторонними источниками).
    env_token = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
    env_chat = (
        os.getenv("WAREHOUSE_REPORT_CHAT_ID", "").strip()
        or os.getenv("TELEGRAM_CHAT_ID", "").strip()
    )
    if env_token and env_chat:
        return env_token, env_chat

    # 2) telegram-bot/.env (локальный продуктовый контур).
    bot_token = env_file.get("TELEGRAM_BOT_TOKEN", "").strip()
    bot_chat = (
        env_file.get("WAREHOUSE_REPORT_CHAT_ID", "").strip()
        or env_file.get("TELEGRAM_CHAT_ID", "").strip()
    )
    if bot_token and bot_chat:
        return bot_token, bot_chat

    # 3) ~/.config/aist/env (каноничный инженерный env).
    aist_token = aist_env.get("TELEGRAM_BOT_TOKEN", "").strip()
    aist_chat = (
        aist_env.get("WAREHOUSE_REPORT_CHAT_ID", "").strip()
        or aist_env.get("TELEGRAM_CHAT_ID", "").strip()
    )
    if aist_token and aist_chat:
        return aist_token, aist_chat

    # 4) ~/.config/exocortex/* (legacy fallback, но пара сохраняется из одного слоя).
    if exocortex_token and exocortex_chat:
        return exocortex_token, exocortex_chat

    # 5) Последний fallback на случай частично заполненного окружения.
    token = env_token or bot_token or aist_token or exocortex_token
    chat_id = env_chat or bot_chat or aist_chat or exocortex_chat
    return token, chat_id


def send_telegram_message(text: str) -> tuple[bool, str]:
    token, chat_id = resolve_telegram_credentials()
    if not token:
        return False, "skip: TELEGRAM_BOT_TOKEN missing"
    if not chat_id:
        return False, "skip: WAREHOUSE_REPORT_CHAT_ID/TELEGRAM_CHAT_ID missing"

    payload = urlencode(
        {
            "chat_id": chat_id,
            "text": text,
            "parse_mode": "HTML",
            "disable_web_page_preview": "true",
        }
    ).encode("utf-8")
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    req = Request(url, data=payload, method="POST")
    try:
        with urlopen(req, timeout=20) as resp:
            ok = 200 <= resp.status < 300
        return ok, f"sent:{ok}"
    except Exception as exc:  # pragma: no cover
        detail = str(exc)
        try:
            detail = f"{detail}; body={exc.read().decode('utf-8', errors='ignore')}"
        except Exception:
            pass
        return False, f"send_failed:{detail}"


def telegram_text(
    cards: list[Path],
    hours: int,
    run_stats: dict[str, int],
    insights: list[dict],
    manual_run: bool,
) -> str:
    ts = datetime.now().strftime("%d.%m.%Y %H:%M")
    report_rel = LATEST_REPORT.relative_to(ROOT).as_posix()
    report_url = build_github_link(report_rel)
    queue_rel = DECISION_QUEUE.relative_to(ROOT).as_posix()
    queue_url = build_github_link(queue_rel)
    procurement_rel = PROCUREMENT_REPORT_FILE.relative_to(ROOT).as_posix()
    procurement_url = build_github_link(procurement_rel) if PROCUREMENT_REPORT_FILE.exists() else ""

    mode = "manual" if manual_run else "auto"
    analytics = build_smart_analytics(insights)
    price_delta = price_delta_from_invoices()
    if (not price_delta.get("up")) and (not price_delta.get("down")):
        price_delta = price_delta_from_catalog(insights)
    sales_top: list[str] = []
    sales_weak: list[str] = []
    for item in insights:
        if item.get("report_type") == "sales":
            sm = item.get("sales_metrics") or {}
            for name, qty, rev in sm.get("top", [])[:3]:
                sales_top.append(f"{name}: {qty:.0f} шт / {rev:,.0f} ₽".replace(",", " "))
            for name, qty, rev in sm.get("weak", [])[:3]:
                sales_weak.append(f"{name}: {qty:.0f} шт / {rev:,.0f} ₽".replace(",", " "))
    sales_top = list(dict.fromkeys(sales_top))[:5]
    sales_weak = list(dict.fromkeys(sales_weak))[:5]
    consumption = consumption_alerts(insights)
    critical_orders = analytics.get("critical_orders") or []
    planned_orders = analytics.get("planned_orders") or []
    reduce_items = analytics.get("reduce_or_stop") or []
    manual_review = list(analytics.get("manual_review") or [])
    no_abc = analytics.get("no_abc") or []
    for item in no_abc[:5]:
        manual_review.append(f"{item}: нет ABC-категории")
    manual_review = list(dict.fromkeys(manual_review))[:6]
    data_risks = detect_data_gaps(insights)

    if run_stats["error"] > 0 or manual_review:
        verdict = "🔴 Требует внимания"
    elif critical_orders:
        verdict = "🟡 Есть срочный заказ"
    else:
        verdict = "🟢 Стабильно"
    executive = executive_summary_lines(
        critical_orders=critical_orders,
        planned_orders=planned_orders,
        reduce_or_stop=reduce_items,
        manual_review=manual_review,
        price_delta=price_delta,
        consumption=consumption,
        has_abc=bool(analytics.get("has_abc")),
    )

    lines = [
        "📦 <b>Склад · управленческий отчёт</b>",
        f"Время: <code>{ts}</code> · Режим: <b>{mode}</b> · Окно: <b>{hours}ч</b>",
        f"Вердикт: <b>{verdict}</b>",
        f"Итог: {' '.join(escape_html(x) for x in executive)}",
        "",
        "<b>Срочно заказать</b>",
    ]

    if critical_orders:
        for supplier_name, items in bucket_orders_by_supplier(critical_orders[:8]):
            channel = str(items[0].get("order_channel", "")).strip()
            contact = str(items[0].get("supplier_contact", "")).strip()
            channel_text = format_supplier_channel(channel, contact)
            deadline = str(items[0].get("deadline", "")).strip() or "уточнить"
            lines.append("• " + escape_html(supplier_name))
            lines.append("  " + escape_html(f"Канал заказа: {channel_text}"))
            lines.append("  " + escape_html(f"Базовый дедлайн: {deadline}"))
            for item in items:
                item_deadline = str(item.get("deadline", "")).strip() or deadline
                extra_deadline = ""
                if item_deadline != deadline:
                    extra_deadline = f"; дедлайн {item_deadline}"
                lines.append(
                    "  - "
                    + escape_html(
                        f"{item['name']}: сейчас {item['qty_now']} шт, заказать {item['qty_to_order']} шт; риск: {item['risk']}{extra_deadline}"
                    )
                )
    else:
        lines.append("• Критичных позиций для заказа нет.")

    lines.append("")
    lines.append("<b>Планово заказать</b>")
    if planned_orders:
        for supplier_name, items in bucket_orders_by_supplier(planned_orders[:8]):
            deadline = str(items[0].get("deadline", "")).strip() or "уточнить"
            lines.append("• " + escape_html(supplier_name))
            lines.append("  " + escape_html(f"Базовый дедлайн: {deadline}"))
            for item in items:
                item_deadline = str(item.get("deadline", "")).strip() or deadline
                extra_deadline = ""
                if item_deadline != deadline:
                    extra_deadline = f"; дедлайн {item_deadline}"
                lines.append(
                    "  - "
                    + escape_html(
                        f"{item['name']}: сейчас {item['qty_now']} шт, дозаказать {item['qty_to_order']} шт; причина: {item['reason']}{extra_deadline}"
                    )
                )
    else:
        lines.append("• Плановых дозаказов не выявлено.")

    lines.append("")
    lines.append("<b>Не заказывать / сократить</b>")
    if reduce_items:
        for item in reduce_items[:5]:
            lines.append(f"• {escape_html(str(item))}")
    else:
        lines.append("• Явных позиций под сокращение закупки нет.")

    lines.append("")
    lines.append("<b>Продажи и ассортимент</b>")
    if sales_top:
        lines.append("• Топ продаж:")
        for item in sales_top[:3]:
            lines.append(f"• {escape_html(item)}")
    if sales_weak:
        lines.append("• Слабые позиции:")
        for item in sales_weak[:3]:
            lines.append(f"• {escape_html(item)}")
    if (not sales_top) and (not sales_weak):
        lines.append("• Недостаточно данных, чтобы уверенно показать топ и слабые позиции.")

    lines.append("")
    lines.append("<b>Цены и расход</b>")
    if price_delta.get("up"):
        for item in price_delta.get("up", [])[:3]:
            lines.append(f"• Рост цены: {escape_html(item)}")
    if price_delta.get("down"):
        for item in price_delta.get("down", [])[:3]:
            lines.append(f"• Снижение цены: {escape_html(item)}")
    if consumption:
        for item in consumption[:3]:
            lines.append(f"• Расход: {escape_html(item)}")
    if (not price_delta.get("up")) and (not price_delta.get("down")) and (not consumption):
        lines.append("• Сильных сигналов по ценам и расходу пока нет.")

    lines.append("")
    lines.append("<b>Проверить вручную</b>")
    if manual_review:
        for item in manual_review[:5]:
            lines.append(f"• {escape_html(str(item))}")
    else:
        lines.append("• Критичных ручных проверок нет.")

    lines.append("")
    lines.append("<b>Каких данных не хватает</b>")
    for item in data_risks[:4]:
        lines.append(f"• {escape_html(item)}")

    if report_url:
        lines.append("")
        lines.append(f"Полный отчёт: <a href=\"{report_url}\">WH.REPORT.002</a>")
    else:
        lines.append("")
        lines.append(f"Полный отчёт: <code>{escape_html(report_rel)}</code>")
    if procurement_url:
        lines.append(f"Закупочный контур: <a href=\"{procurement_url}\">WH.WP.007 supplier-map</a>")
    elif PROCUREMENT_REPORT_FILE.exists():
        lines.append(f"Закупочный контур: <code>{escape_html(procurement_rel)}</code>")
    if queue_url:
        lines.append(f"Очередь решений: <a href=\"{queue_url}\">WH.SESSION.001</a>")
    else:
        lines.append(f"Очередь решений: <code>{escape_html(queue_rel)}</code>")

    return "\n".join(lines)


def refresh_procurement_report() -> tuple[bool, str]:
    script = WAREHOUSE_DIR / "tools" / "warehouse_invoice_procurement_report.py"
    if not script.exists():
        return False, "script_missing"
    try:
        proc = subprocess.run(
            ["python3", str(script)],
            cwd=str(ROOT),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=240,
            check=False,
        )
        if proc.returncode == 0:
            return True, (proc.stdout or "").strip()[:240]
        return False, ((proc.stderr or proc.stdout or "").strip()[:240] or f"exit={proc.returncode}")
    except Exception as exc:
        return False, str(exc)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--hours", type=int, default=6)
    parser.add_argument("--send-telegram", action="store_true")
    parser.add_argument("--telegram-on-empty", action="store_true")
    parser.add_argument("--manual-run", action="store_true")
    parser.add_argument("--refresh-duplicates", action="store_true")
    parser.add_argument("--replay-latest-cards", type=int, default=0)
    args = parser.parse_args()

    WP_DIR.mkdir(parents=True, exist_ok=True)
    CARDS_DIR.mkdir(parents=True, exist_ok=True)
    DLQ_DIR.mkdir(parents=True, exist_ok=True)

    sources = find_recent_reports(hours=args.hours)
    registry = load_registry()
    stamp = now_stamp()
    run_stats = {"received": len(sources), "processed": 0, "duplicate": 0, "error": 0, "dlq": 0}
    cards: list[Path] = []
    bot_cards: list[Path] = []
    insights: list[dict] = []
    for src in sources:
        rel = src.relative_to(ROOT).as_posix()
        src_stat = src.stat()
        src_mtime = datetime.fromtimestamp(src_stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        src_hash = file_sha1(src)
        record_key = f"{rel}|{src_stat.st_size}|{int(src_stat.st_mtime)}|{src_hash[:12]}"

        existing = registry.get(record_key)
        if existing:
            if args.refresh_duplicates:
                try:
                    card, bot_card, insight = make_card(src)
                    cards.append(card)
                    bot_cards.append(bot_card)
                    insights.append(insight)
                    existing["status"] = "refreshed"
                    existing["last_seen_at"] = stamp
                    existing["last_processed_at"] = stamp
                    existing["card_path"] = card.relative_to(ROOT).as_posix()
                    existing["bot_card_path"] = bot_card.relative_to(ROOT).as_posix()
                    existing["error"] = ""
                    run_stats["processed"] += 1
                except Exception as exc:  # pragma: no cover
                    existing["status"] = "error"
                    existing["last_seen_at"] = stamp
                    existing["error"] = str(exc)
                    run_stats["error"] += 1
            else:
                run_stats["duplicate"] += 1
                existing["status"] = "duplicate"
                existing["last_seen_at"] = stamp
            continue

        row = {
            "record_key": record_key,
            "source_file": rel,
            "source_mtime": src_mtime,
            "source_size_bytes": str(src_stat.st_size),
            "source_hash": src_hash,
            "report_type": detect_report_type(src.name),
            "status": "new",
            "first_seen_at": stamp,
            "last_seen_at": stamp,
            "last_processed_at": "",
            "card_path": "",
            "bot_card_path": "",
            "dlq_path": "",
            "error": "",
        }

        try:
            card, bot_card, insight = make_card(src)
            cards.append(card)
            bot_cards.append(bot_card)
            insights.append(insight)
            row["status"] = "processed"
            row["last_processed_at"] = stamp
            row["card_path"] = card.relative_to(ROOT).as_posix()
            row["bot_card_path"] = bot_card.relative_to(ROOT).as_posix()
            run_stats["processed"] += 1
        except Exception as exc:  # pragma: no cover
            row["status"] = "error"
            row["error"] = str(exc)[:500]
            row["dlq_path"] = append_dlq_entry(src, record_key, row["error"])
            run_stats["error"] += 1
            run_stats["dlq"] += 1

        registry[record_key] = row

    save_registry(registry)
    refresh_ok, refresh_info = refresh_procurement_report()
    if refresh_ok:
        print(f"[warehouse] procurement_refresh=ok {refresh_info}")
    else:
        print(f"[warehouse] procurement_refresh=warn {refresh_info}")
    if not insights and args.manual_run and sources:
        replay_cards, replay_bot_cards, replay_insights = rebuild_insights_from_sources(sources, limit=24)
        cards.extend(replay_cards)
        bot_cards.extend(replay_bot_cards)
        insights = replay_insights
    elif not insights and args.replay_latest_cards > 0:
        insights = replay_insights_from_registry(registry, limit=args.replay_latest_cards)

    build_latest_summary(cards, bot_cards, args.hours, run_stats, insights)
    build_decision_queue(insights, run_stats, args.manual_run)
    print(
        f"[warehouse] sources={len(sources)} cards={len(set(cards))} "
        f"bot_cards={len(set(bot_cards))} duplicates={run_stats['duplicate']} "
        f"errors={run_stats['error']} summary={LATEST_REPORT} registry={REGISTRY_FILE}"
    )

    if args.send_telegram:
        if (not args.manual_run) and (os.getenv("WAREHOUSE_ALLOW_AUTO_TELEGRAM", "0").strip() != "1"):
            print("[warehouse] telegram=skip:auto-disabled (use --manual-run or WAREHOUSE_ALLOW_AUTO_TELEGRAM=1)")
            return 0
        if (not cards) and (not args.telegram_on_empty):
            print("[warehouse] telegram=skip:no-new-cards")
            return 0
        text = telegram_text(cards, args.hours, run_stats, insights, args.manual_run)
        ok, info = send_telegram_message(text)
        print(f"[warehouse] telegram={info}")
        if not ok:
            # Не валим пайплайн из-за Telegram; фиксируем как warning.
            return 0
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
